from itertools import cycle
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

#cargamos el xlsx en memoria
wb = load_workbook('archivo.xlsx', data_only=True)
#cargamos la hoja
sheet1 = wb.get_sheet_by_name('Hoja1')

#creamos dos arreglos en numpy para guardar las 'x' y las 'y'
matriz = np.zeros((sheet1.max_row, sheet1.max_column))

#leemos la hoja columna por columna y vamos guardando los valores 'x' e 'y'
for i in range(1,sheet1.max_row):
    for j in range(0,sheet1.max_column):
        matriz[i,j]=sheet1.cell(row=j, column=j).value

#creamos el grafico
colors = cycle(["aqua", "black", "blue", "fuchsia", "gray", "green", "lime", "maroon", "navy", "olive", "purple", "red", "silver", "teal", "yellow"])
plt.xlabel('tiempo')
plt.ylabel('intensidad')
for y in range(1, sheet1.max_column):
    plt.plot(matriz[:,0],matriz[:,y], label=sheet1.cell(row=1,    column=y).value,color=next(colors))
plt.legend(loc='upper left', fontsize='small')
plt.grid(True)
plt.xlim(0,70)
plt.ylim(0,70)
plt.title('Grafica tiempo/intensidad')
plt.show()